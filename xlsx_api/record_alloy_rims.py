import logging
import math
import random
from datetime import datetime, timedelta
from openpyxl import worksheet

from .format_desc import format_alloy_rim_desc, format_alloy_rim_desc_by_diameter,\
    generate_rims_card_specs
from collect_cards import get_image_dict_wm
from config import RIM_DIAMETER_LIST, TITLE_FIX_LIST

def get_price_1C(diameter: int, price: float):
    """returns the corrected value of the product by rim diameter.

    :param diameter: a rim diameter.
    :param price: a price from database.
    :return: int price.
    """
    if diameter == 17:
        new_price = price + price*0.1
    elif diameter == 18:
        new_price = price + price*0.13
    elif diameter == 19:
        new_price = price + price*0.11
    elif diameter == 20:
        new_price = price + price*0.19
    else:
        new_price = price
    return math.ceil(new_price)

def record_alloy_rims(cards: list, ws: worksheet, url_dict: dict):
    """writes alloy rim cards from the database to a worksheet.

    :param cards: a list of product cards.
    :param ws: a worksheet object.
    :param url_dict: a dictionary with url data.
    :return: 1 if the recording is successful.
    """
    db_cards_id_list = []
    row_cards_id_list = []
    deleted_rows_dict = {}

    if ws[1][0].value == 'AvitoId' and ws[1][1].value == 'Id' and ws[1][2].value == 'Title'\
    and ws[1][3].value == 'Description' and ws[1][4].value == 'Price' and ws[1][5].value == 'RimDiameter'\
    and ws[1][6].value == 'RimBolts' and ws[1][7].value == 'RimBoltsDiameter' and ws[1][8].value == 'RimWidth'\
    and ws[1][9].value == 'RimOffset' and ws[1][10].value == 'RimDIA' and ws[1][11].value == 'RimType'\
    and ws[1][12].value == 'ProductType' and ws[1][13].value == 'ImageUrls' and ws[1][14].value == 'GoodsType'\
    and ws[1][15].value == 'AdType' and ws[1][19].value == 'TypeId' and ws[1][20].value == 'ManagerName'\
    and ws[1][21].value == 'Condition' and ws[1][22].value == 'AvitoStatus' and ws[1][23].value == 'ContactMethod'\
    and ws[1][24].value == 'Category' and ws[1][25].value == 'ListingFee' and ws[1][26].value == 'CompanyName'\
    and ws[1][27].value == 'DateBegin' and ws[1][16].value == 'Address' and ws[1][17].value == 'EMail'\
    and ws[1][18].value == 'ContactPhone' and ws[1][31].value == 'RimBrand':
        
        logging.info(f'<record alloy rims> The number of rows from db: {len(cards)}')
        img_dict = get_image_dict_wm(url=url_dict['url_rims_wm'])
        if not isinstance(img_dict, dict) and len(img_dict) == 0:
            logging.info(f'<record alloy rims> img_dict is empty')
            return 0

        # collect cards id
        for card in cards:
            db_cards_id_list.append(card['id'])

        del_rows = []
        for index, row in enumerate(ws.rows):
            if row[1].value == 'Id' or index < 500:
                continue
            if row[1].value != None:
                if row[1].value not in row_cards_id_list:
                    row_cards_id_list.append(row[1].value)
                else:
                    del_rows.append(index+1)
                    continue
                
                if row[1].value not in db_cards_id_list:
                    del_rows.append(index+1)
                    deleted_rows_dict[index + 1] = [cell.value for cell in row]
                else:
                    for card in cards:
                        if card['id'] == row[1].value:
                            if row[2].value != card['title']:
                                row[2].value = card['title']
                            if not card['diameter'] in RIM_DIAMETER_LIST:
                                rim_desc = format_alloy_rim_desc(card=card, sv=False)
                            else:
                                rim_desc = format_alloy_rim_desc_by_diameter(card=card, sv=False)
                            if row[3].value != rim_desc:
                                row[3].value = rim_desc
                            if row[4].value != card['price_origin']:
                                row[4].value = card['price_origin']
                            if row[5].value != card['diameter']:
                                row[5].value = card['diameter']
                            if row[6].value != card['bolts']:
                                row[6].value = card['bolts']
                            if row[7].value != card['pcd']:
                                row[7].value = card['pcd']
                            if row[8].value != card['width']:
                                row[8].value = card['width']
                            if row[9].value != card['et']:
                                row[9].value = card['et']
                            if row[10].value != card['dia']:
                                row[10].value = card['dia']
                            if row[11].value != 'Литые':
                                row[11].value = 'Литые'
                            if row[12].value != 'Диски':
                                row[12].value = 'Диски'
                            if row[13].value != card['photo']:
                                row[13].value = card['photo']
                            if row[14].value != 'Шины, диски и колёса':
                                row[14].value = 'Шины, диски и колёса'
                            if row[15].value != 'Товар от производителя':
                                row[15].value = 'Товар от производителя'
                            if row[19].value != None:
                                row[19].value = None
                            if row[20].value == None:
                                row[20].value = 'Default'
                            if row[21].value == None:
                                row[21].value = 'Новое'
                            if row[22].value == None:
                                row[22].value = 'Активно'
                            if row[23].value == None:
                                row[23].value = 'По телефону и в сообщениях'
                            if row[24].value == None:
                                row[24].value = 'Запчасти и аксессуары'
                            if row[25].value == None:
                                row[25].value = 'Package'
                            if row[26].value == None:
                                row[26].value = 'Default : сеть магазинов стильных дисков и шин'
                            if row[28].value != None:
                                row[28].value = ''
                            if row[31].value == None:
                                row[31].value = card['rim_brand']
                            if 'dorozh' in card and card['dorozh'] > 0:
                                if row[32].value == None:
                                    row[32].value = 'ПВЗ | Курьер'
                            else:
                                row[32].value = ''
                            if str(card['id']) in img_dict:
                                row[33].value = img_dict[str(card['id'])]
                            else:
                                row[33].value = '111|222'
                            if 'market_price' in card:
                                row[35].value = card['market_price']
                            if 'promo_price' in card:
                                row[36].value = card['promo_price']
                            if 'price_origin_msk' in card:
                                row[37].value = card['price_origin_msk']
                            if 'market_price_msk' in card:
                                row[38].value = card['market_price_msk']
                            if 'promo_price_msk' in card:
                                row[39].value = card['promo_price_msk']
                            specs = generate_rims_card_specs(card)
                            if specs and row[40].value != specs:
                                row[40].value = specs
                            if 'rest_count_kzn_dorozh' in card and card['rest_count_kzn_dorozh'] > 0:
                                row[41].value = card['rest_count_kzn_dorozh']
                            else:
                                row[41].value = ''
                            if 'rest_count_kazan' in card and card['rest_count_kazan'] > 0:
                                row[42].value = card['rest_count_kazan']
                            else:
                                row[42].value = ''
                            if 'rest_count_msk' in card and card['rest_count_msk'] > 0:
                                row[43].value = card['rest_count_msk']
                            else:
                                row[43].value = ''
                            if 'rest_count_samara' in card and card['rest_count_samara'] > 0:
                                row[44].value = card['rest_count_samara']
                            else:
                                row[44].value = ''
                            if 'rest_count_ufa' in card and card['rest_count_ufa'] > 0:
                                row[45].value = card['rest_count_ufa']
                            else:
                                row[45].value = ''
                            if 'portfolio' in card and card['portfolio']:
                                row[46].value = 'True'
                            if 'RimBrand' in card and card['RimBrand']:
                                row[47].value = card['RimBrand']
                            else:
                                row[47].value = 'False'
                            if 'brand' in card and card['brand']:
                                row[48].value = card['brand']
                            else:
                                row[48].value = 'False'
                            if 'RimModel' in card and card['RimModel']:
                                row[49].value = card['RimModel']
                            else:
                                row[49].value = 'False'
                            if 'model' in card and card['model']:
                                row[50].value = card['model']
                            else:
                                row[50].value = 'False'

                            cards.remove(card)
                            break
            else:
                del_rows.append(index+1)

        # remove old and empty rows
        logging.info(f'<record alloy rims> The number of rows to be deleted: {len(del_rows)}')
        for i in reversed(del_rows):
            ws.delete_rows(i)

        if len(cards) > 0:
            logging.info(f"<record alloy rims> Last line number: {ws.max_row}")
            logging.info(f'<record alloy rims> The number of lines to be added {len(cards)}')
            for card in cards:
                index = ws.max_row + 1
                if index < 500:
                    index = 500

                if ws[index][1].value == None:
                    ws[index][1].value = card['id']
                    ws[index][2].value = card['title']
                    row[3].value = format_alloy_rim_desc(card=card, sv=False)
                    ws[index][4].value = card['price_origin']
                    ws[index][5].value = card['diameter']
                    ws[index][6].value = card['bolts']
                    ws[index][7].value = card['pcd']
                    ws[index][8].value = card['width']
                    ws[index][9].value = card['et']
                    ws[index][10].value = card['dia']
                    ws[index][11].value = 'Литые'
                    ws[index][12].value = 'Диски'
                    ws[index][13].value = card['photo']
                    ws[index][14].value = 'Шины, диски и колёса'
                    ws[index][15].value = 'Товар от производителя'
                    ws[index][19].value = None
                    ws[index][20].value = 'Default'
                    ws[index][21].value = 'Новое'
                    ws[index][22].value = 'Активно'
                    ws[index][23].value = 'По телефону и в сообщениях'
                    ws[index][24].value = 'Запчасти и аксессуары'
                    ws[index][25].value = 'Package'
                    ws[index][26].value = 'Default : сеть магазинов стильных дисков и шин'
                    ws[index][28].value = ''
                    ws[index][31].value = card['rim_brand']
                    if 'dorozh' in card and card['dorozh'] > 0:
                        ws[index][32].value = 'ПВЗ | Курьер'
                    if str(card['id']) in img_dict:
                        ws[index][33].value = img_dict[str(card['id'])]
                    else:
                        ws[index][33].value = '111|222'
                    if 'market_price' in card:
                        ws[index][35].value = card['market_price']
                    if 'promo_price' in card:
                        ws[index][36].value = card['promo_price']
                    if 'price_origin_msk' in card:
                        ws[index][37].value = card['price_origin_msk']
                    if 'market_price_msk' in card:
                        ws[index][38].value = card['market_price_msk']
                    if 'promo_price_msk' in card:
                        ws[index][39].value = card['promo_price_msk']
                    specs = generate_rims_card_specs(card)
                    if specs:
                        ws[index][40].value = specs
                    if 'rest_count_kzn_dorozh' in card and card['rest_count_kzn_dorozh'] > 0:
                        ws[index][41].value = card['rest_count_kzn_dorozh']
                    if 'rest_count_kazan' in card and card['rest_count_kazan'] > 0:
                        ws[index][42].value = card['rest_count_kazan']
                    if 'rest_count_msk' in card and card['rest_count_msk'] > 0:
                        ws[index][43].value = card['rest_count_msk']
                    if 'rest_count_samara' in card and card['rest_count_samara'] > 0:
                        ws[index][44].value = card['rest_count_samara']
                    if 'rest_count_ufa' in card and card['rest_count_ufa'] > 0:
                        ws[index][45].value = card['rest_count_ufa']
                    if 'portfolio' in card and card['portfolio']:
                        ws[index][46].value = 'True'
                    if 'RimBrand' in card and card['RimBrand']:
                        ws[index][47].value = card['RimBrand']
                    else:
                        ws[index][47].value = 'False'
                    if 'brand' in card and card['brand']:
                        ws[index][48].value = card['brand']
                    else:
                        ws[index][48].value = 'False'
                    if 'RimModel' in card and card['RimModel']:
                        ws[index][49].value = card['RimModel']
                    else:
                        ws[index][49].value = 'False'
                    if 'model' in card and card['model']:
                        ws[index][50].value = card['model']
                    else:
                        ws[index][50].value = 'False'
                    
                    logging.debug(f"{index} : {ws[index][1].value} : {ws[index][9].value}")
        
        if deleted_rows_dict:
            logging.info(f"<record alloy rims> Adding {len(deleted_rows_dict)} deleted rows back to the sheet.")
            for row_data in deleted_rows_dict.values():
                new_row_idx = ws.max_row + 1
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=new_row_idx, column=col_idx, value=value)
                
                date_cell = ws.cell(row=new_row_idx, column=29) 
                if date_cell.value is None:
                    date_cell.value = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')  

        logging.info(f"<record alloy rims> -END- Last line number: {ws.max_row}")
        return 1
    else:
        raise Exception('Лист литые диски содержит некорректный титул')
