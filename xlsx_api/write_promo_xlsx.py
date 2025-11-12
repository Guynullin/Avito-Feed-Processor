import os
import logging
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException
from typing import List, Dict, Union


def write_promo_xlsx(
    save_path: str,
    region: str,
    filter_key: Union[str, List[str]],
    cards_list: List[Dict],
    origin: str = "Price",
    promo: str = "promo_price",
    simple_id: bool = False,
    discount_percent: int = 0
) -> bool:
    """Writes filtered product data to xlsx file.
    
    :param save_path: Full path to directory including filename
    :param region: Region name for filling the region column
    :param filter_key: Key or list of keys for filtering products
    :param cards_list: List of dictionaries with product data
    :param origin: Key name for origin price (default: "origin_price")
    :param promo: Key name for promo price (default: "promo_price")
    :param simple_id: id with RZ and 373 in the end (default: False, with RZ and 373 in the end)
    :param discount_percent: (default: 0)
    :return: True if write successful, False if error occurred
    """

    new_id_list = [733138, 725771, 725758, 725770, 725775, 725744, 
               725777, 725752, 725776, 725779, 725749, 725768, 725751, 725781]

    try:
        directory = os.path.dirname(save_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)
            logging.info(f"Created directory: {directory}")
        
        wb = Workbook()
        
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        ws = wb.create_sheet(title="sale_items")
        
        headers = [
            "ID", "Avito ID", "Название", "Категория", "Микрокатегория",
            "Регион", "Цена", "Скидка, %", "Минимальная, %", "Максимальная, %"
        ]
        ws.append(headers)
        
        if isinstance(filter_key, str):
            filter_keys = [filter_key]
        else:
            filter_keys = filter_key
        
        valid_cards = []
        
        for card in cards_list:
            has_valid_filter_key = False
            for key in filter_keys:
                if (key in card and 
                    card[key] is not None and 
                    str(card[key]).strip() != ''):
                    try:
                        value = int(card[key])
                        if value > 0:
                            has_valid_filter_key = True
                            break
                    except (ValueError, TypeError):
                        continue
            
            if not has_valid_filter_key:
                continue
            
            if 'DateEnd' in card and card['DateEnd'] is not None:
                if (isinstance(card['DateEnd'], str) and 
                    card['DateEnd'].strip() != '') or card['DateEnd']:
                    continue
            
            if discount_percent == 0:
                required_keys = ['Id', 'Title', origin, promo]
            else:
                required_keys = ['Id', 'Title', origin]

            if not all(key in card for key in required_keys):
                continue
            
            if (card['Id'] is None or 
                card['Title'] is None or 
                card[origin] is None):
                continue

            try:
                if discount_percent == 0:
                    origin_price = int(card[origin])
                    promo_price = int(card[promo])
                    
                    if not (origin_price > 0 and promo_price > 0):
                        continue
                    
                    if not (origin_price > promo_price):
                        continue
                    
                    discount_percent = ((origin_price - promo_price) / origin_price) * 100
                    if discount_percent > 35:
                        continue
                else:
                    if not (origin_price > 0):
                        continue
                    
            except (ValueError, TypeError, ZeroDivisionError):
                continue
            
            valid_cards.append(card)
        
        if not simple_id:
            if int(card['Id']) in new_id_list:
                id = f"NZ{card['Id']}373"
            else:
                id = f"RZ{card['Id']}373"
        else:
            id = card['Id']
        
        for card in valid_cards:
            try:
                origin_price = int(card[origin])
                if discount_percent == 0:
                    promo_price = int(card[promo])
                    discount_percent = round(((origin_price - promo_price) / origin_price) * 100)
                
                row = [
                    id,                    
                    "",                           
                    card['Title'],                
                    "Запчасти и аксессуары",     
                    "Диски",                      
                    region,                       
                    origin_price,                 
                    discount_percent,             
                    "",                           
                    ""                            
                ]
                ws.append(row)
                
            except (ValueError, TypeError, KeyError) as e:
                logging.warning(f"Error processing card {card.get('Id', 'unknown')}: {e}")
                continue
        
        wb.save(save_path)
        logging.info(f"Successfully wrote {len(valid_cards)} cards to {save_path}")
        return True
        
    except Exception as e:
        logging.error(f"Error writing xlsx file {save_path}: {e}")
        return False