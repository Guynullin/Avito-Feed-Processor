import os
import copy
import logging
from openpyxl import load_workbook

from .record_forged_rims import record_forged_rims
from .record_alloy_rims import record_alloy_rims
from .record_tires import record_tires
from info_bot import send_message

def record_xlsx(input_file_path: str, db_cards: dict, input_filename: str, url_dict: dict):
    """writes data from the database to a xlx file.

    :param input_file_path: a path to the input file directory .
    :param db_cards: a cards from database.
    :param input_filename: a name of input file.
    :param url_dict: a dictionary with url data.
    :return: 1 if the recording is successful, or 0 if error occured. 
    """
    flag = 1

    for root, dirs, files in os.walk(input_file_path):  
        for filename in files:
            if filename == input_filename:
                file_path = os.path.join(input_file_path, filename)
                if os.path.exists(file_path) and os.path.isfile(file_path):
                    wb = load_workbook(file_path)
                else:
                    send_message(f'ERROR\nfile path: {file_path} not exists')
                    logging.error(f'<record_xlsx> file path: {file_path} not exists')
                    return 0
                cards = copy.deepcopy(db_cards)
                for sheetname in wb.sheetnames:
                    ws = wb[sheetname]
                    if sheetname == 'диски кованые':
                        if 'forged' in cards:
                            if record_forged_rims(cards=cards['forged'], ws=ws, url_dict=url_dict) != 1:
                                logging.error('ERROR\nЛист кованые диски не записан!')
                                send_message('ERROR\nЛист кованые диски не записан!')
                                flag = 0
                        else:
                            logging.error('ERROR\nЛист кованые диски не записан!\nВ выгрузке нет кованых дисков')
                            send_message('ERROR\nЛист кованые диски не записан!\nВ выгрузке нет кованых дисков')
                            flag = 0
                    elif sheetname == 'диски литые':
                        if 'alloy' in cards:
                            if record_alloy_rims(cards=cards['alloy'], ws=ws, url_dict=url_dict) != 1:
                                logging.error('ERROR\nЛист литые диски не записан!')
                                send_message('ERROR\nЛист литые диски не записан!')
                                flag = 0
                        else:
                            logging.error('ERROR\nЛист литые диски не записан!\nВ выгрузке нет литых дисков')
                            send_message('ERROR\nЛист литые диски не записан!\nВ выгрузке нет литых дисков')
                            flag = 0
                    # Deprecated
                    # elif sheetname == 'шины':
                    #     if 'tires' in cards:
                    #         if record_tires(cards_list=cards['tires'], wb=wb, url_dict=url_dict) != 1:
                    #                 logging.error('ERROR\nЛист шины не записан!')
                    #                 send_message('ERROR\nЛист шины не записан!')
                    #                 flag = 0
                    #     else:
                    #         logging.error('ERROR\nЛист шины не записан!\nВ выгрузке нет шин')
                    #         send_message('ERROR\nЛист шины не записан!\nВ выгрузке нет шин')
                    #         flag = 0

                    # Deprecated
                    # elif sheetname == 'пружины':
                    #     if record_springs(cards=cards['springs'], ws=ws, city_data=CITY_DATA[filename]) != 1:
                    #         send_message('ERROR\nЛист пружины не записан!')

                wb.save(file_path)
                wb.close()
                logging.info(f'<record_xlsx> the file was saved successfully. Path: {file_path}')

    return flag
