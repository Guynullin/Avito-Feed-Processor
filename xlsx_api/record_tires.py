import logging

from openpyxl import workbook
from .format_desc import format_tire_desc
from collect_cards import get_image_dict_wm
from config import TIRES_VIDEO_URL
from datetime import datetime, timedelta


def record_tires(cards_list: list, wb: workbook, url_dict: dict):
    """writes tire cards from the database to a worksheet.

    :param cards_list: a list of product cards.
    :param wb: a workbook object.
    :param url_dict: a dictionary with url data.
    :return: 1 if the recording is successful.
    """
    ws = wb['шины']
    db_cards_id_list = []
    row_cards_id_list = []
    deleted_tires_dict = {}


    if ws[1][0].value == 'AvitoId' and ws[1][1].value == 'Id' and ws[1][2].value == 'Title'\
        and ws[1][3].value == 'Description' and ws[1][4].value == 'Price' and ws[1][5].value == 'RimDiameter'\
        and ws[1][6].value == 'ProductType' and ws[1][7].value == 'ImageUrls' and ws[1][8].value == 'GoodsType'\
        and ws[1][9].value == 'AdType' and ws[1][10].value == 'Address' and ws[1][11].value == 'EMail'\
        and ws[1][12].value == 'ContactPhone' and ws[1][13].value == 'TypeId' and ws[1][14].value == 'Condition'\
        and ws[1][15].value == 'AvitoStatus' and ws[1][16].value == 'ContactMethod' and ws[1][17].value == 'Category'\
        and ws[1][18].value == 'ListingFee' and ws[1][19].value == 'CompanyName' and ws[1][20].value == 'Quantity'\
        and ws[1][21].value == 'TireType' and ws[1][22].value == 'TireAspectRatio' and ws[1][23].value == 'LoadIndex'\
        and ws[1][24].value == 'DifferentWidthTires' and ws[1][25].value == 'SpeedIndex' and ws[1][26].value == 'RunFlat'\
        and ws[1][27].value == 'Model' and ws[1][28].value == 'Brand' and ws[1][29].value == 'TireSectionWidth'\
        and ws[1][30].value == 'VideoURL':

        logging.info(f'<record tires> The number of rows from db: {len(cards_list)}')
        img_dict = get_image_dict_wm(url=url_dict['url_tires_wm'])
        if not isinstance(img_dict, dict) and len(img_dict) == 0:
            logging.info(f'<record alloy rims> img_dict is empty')
            return 0
        # collect cards id
        for card in cards_list:
            db_cards_id_list.append(card['id'])

        del_rows = []
        for index, row in enumerate(ws.rows):
            if row[1].value == 'Id':
                continue
            if row[1].value != None:
                if not row[1].value in row_cards_id_list:
                    row_cards_id_list.append(row[1].value)
                else:
                    del_rows.append(index+1)
                    continue
                
                if not row[1].value in db_cards_id_list:
                    del_rows.append(index+1)
                    deleted_tires_dict[index + 1] = [cell.value for cell in row]
                else:
                    for card in cards_list:
                        if card['id'] == row[1].value:
                            if row[2].value != card['title'].replace('*', '-'):
                                row[2].value = card['title'].replace('*', '-')
                            # tire_desc = format_tire_desc(season=card['tire_type'], title=card['title'])
                            # if row[3].value != tire_desc:
                            #     row[3].value = tire_desc
                            if row[4].value != card['price_origin']:
                                row[4].value = card['price_origin']
                            if row[5].value != card['diameter']:
                                row[5].value = card['diameter']
                            if row[6].value != 'Легковые шины':
                                row[6].value = 'Легковые шины'
                            if row[7].value != card["photo"]:
                                row[7].value = card["photo"]
                            if row[8].value != 'Шины, диски и колёса':
                                row[8].value = 'Шины, диски и колёса'
                            if row[9].value != 'Товар от производителя':
                                row[9].value = 'Товар от производителя'
                            if row[13].value != None:
                                row[13].value = None
                            if row[14].value != 'Новое':
                                row[14].value = 'Новое'
                            if row[15].value != None:
                                if row[15].value != 'Активно':
                                    row[15].value = 'Активно'
                            if row[16].value != 'По телефону и в сообщениях':
                                row[16].value = 'По телефону и в сообщениях'
                            if row[17].value != 'Запчасти и аксессуары':
                                row[17].value = 'Запчасти и аксессуары'
                            if row[18].value != 'Package':
                                row[18].value = 'Package'
                            if row[19].value != 'Default : сеть магазинов стильных дисков и шин':
                                row[19].value = 'Default : сеть магазинов стильных дисков и шин'
                            if row[20].value != 'за 1 шт.':
                                row[20].value = 'за 1 шт.'
                            if row[21].value != card['tire_type']:
                                row[21].value = card['tire_type']
                            if row[22].value != card['height']:
                                row[22].value = card['height']
                            if row[23].value != card['load_index']:
                                row[23].value = card['load_index']
                            if row[24].value != 'Нет':
                                row[24].value = 'Нет'
                            if row[25].value != card['speed_index']:
                                row[25].value = card['speed_index']
                            if row[26].value != card['runflat']:
                                row[26].value = card['runflat']
                            if row[27].value != card['model']:
                                row[27].value = card['model']
                            if row[28].value != card['brand']:
                                row[28].value = card['brand']
                            if row[29].value != card['width']:
                                row[29].value = card['width']
                            if row[30].value != TIRES_VIDEO_URL:
                                row[30].value = TIRES_VIDEO_URL
                            if row[32].value != None:
                                row[32].value = ''
                            if str(card['id']) in img_dict:
                                row[36].value = img_dict[str(card['id'])]
                            else:
                                # row[36].value = card['main_img']
                                row[36].value = '111|222'
                            cards_list.remove(card)
                            break
            else:
                del_rows.append(index+1)

        # remove old and empty rows
        logging.info(f'<record_tires> The number of rows to be deleted: {len(del_rows)}')
        for i in reversed(del_rows):
            ws.delete_rows(i)

        if len(cards_list) > 0:
            logging.info(f"<record_tires> Last line number: {ws.max_row}")
            logging.info(f'<record_tires> The number of lines to be added {len(cards_list)}')
            for card in cards_list:
                index = ws.max_row + 1

                ws[index][1].value = card['id']
                ws[index][2].value = card['title'].replace('*', '-')
                # ws[index][3].value = format_tire_desc(season=card['tire_type'], title=card['title'])
                ws[index][4].value = card['price_origin']
                ws[index][5].value = card['diameter']
                ws[index][6].value = 'Легковые шины'
                ws[index][7].value = card["photo"]
                ws[index][8].value = 'Шины, диски и колёса'
                ws[index][9].value = 'Товар от производителя'
                ws[index][13].value = None
                ws[index][14].value = 'Новое'
                ws[index][15].value = 'Активно'
                ws[index][16].value = 'По телефону и в сообщениях'
                ws[index][17].value = 'Запчасти и аксессуары'
                ws[index][18].value = 'Package'
                ws[index][19].value = 'Default : сеть магазинов стильных дисков и шин'
                ws[index][20].value = 'за 1 шт.'
                ws[index][21].value = card['tire_type']
                ws[index][22].value = card['height']
                ws[index][23].value = card['load_index']
                ws[index][24].value = 'Нет'
                ws[index][25].value = card['speed_index']
                ws[index][26].value = card['runflat']
                ws[index][27].value = card['model']
                ws[index][28].value = card['brand']
                ws[index][29].value = card['width']
                ws[index][30].value = TIRES_VIDEO_URL
                ws[index][32].value = ''
                if str(card['id']) in img_dict:
                        ws[index][36].value = img_dict[str(card['id'])]
                else:
                    # ws[index][36].value = card['main_img']
                    ws[index][36].value = '111|222'
                logging.debug(f"{index} : {ws[index][1].value} : {ws[index][9].value}")
        
        if deleted_tires_dict:
            logging.info(f"<record_tires> Adding {len(deleted_tires_dict)} deleted rows back to the sheet.")
            for row_data in deleted_tires_dict.values():
                new_row_idx = ws.max_row + 1
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=new_row_idx, column=col_idx, value=value)
                
                date_cell = ws.cell(row=new_row_idx, column=33) 
                if date_cell.value is None:
                    date_cell.value = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')  

        logging.info(f"<record_tires> -END- Last line number: {ws.max_row}")
        return 1
    else:
        raise Exception('Лист шины содержит некорректный титул')
