from openpyxl import worksheet
import logging

def parse_tires(ws: worksheet):
    """returns a list with tire cards collected from the xlsx table.

    :param ws: a worksheet object.
    :return: a list with the cards or zero when error occured.
    """
    cards_list = []

    if ws[1][0].value == 'AvitoId' and ws[1][1].value == 'Id' and ws[1][2].value == 'Title'\
        and ws[1][3].value == 'Description' and ws[1][4].value == 'Price' and ws[1][5].value == 'RimDiameter'\
        and ws[1][6].value == 'ProductType' and ws[1][7].value == 'ImageUrls' and ws[1][8].value == 'GoodsType'\
        and ws[1][9].value == 'AdType' and ws[1][10].value == 'Address' and ws[1][11].value == 'EMail'\
        and ws[1][12].value == 'ContactPhone' and ws[1][13].value == 'TypeId' and ws[1][14].value == 'Condition'\
        and ws[1][15].value == 'AvitoStatus' and ws[1][16].value == 'ContactMethod' and ws[1][17].value == 'Category'\
        and ws[1][18].value == 'ListingFee' and ws[1][19].value == 'CompanyName' and ws[1][20].value == 'Quantity'\
        and ws[1][21].value == 'TireType' and ws[1][22].value == 'TireAspectRatio' and ws[1][23].value == 'LoadIndex'\
        and ws[1][24].value == 'DifferentWidthTires' and ws[1][25].value == 'SpeedIndex' and ws[1][26].value == 'RunFlat'\
        and ws[1][27].value == 'Model' and ws[1][28].value == 'Brand' and ws[1][29].value == 'TireSectionWidth'\
        and ws[1][30].value == 'VideoURL' and ws[1][31].value == 'DateBegin':
        

        for index, row in enumerate(ws.rows):
            if row[1].value == 'Id':
                continue

            if row[1].value != None and index <= ws.max_row:
                card = {"Id" : row[1].value}

                if row[2].value != None:
                    card['Title'] = row[2].value
                else:
                    continue
                if row[3].value != None:
                    card['Description'] = row[3].value
                else:
                    continue
                if row[4].value != None:
                    card['Price'] = row[4].value
                else:
                    continue 
                if row[5].value != None:
                    card['RimDiameter'] = row[5].value
                else:
                    continue
                if row[6].value != None:
                    card['ProductType'] = row[6].value
                else:
                    continue
                if row[7].value != None:
                    card['ImageUrls'] = row[7].value
                else:
                    continue
                if row[8].value != None:
                    card['GoodsType'] = row[8].value
                else:
                    continue
                if row[9].value != None:
                    card['AdType'] = row[9].value
                else:
                    continue
                if row[13].value != None:
                    card['TypeId'] = row[13].value
                if row[14].value != None:
                    card['Condition'] = row[14].value
                else:
                    continue
                if row[15].value != None:
                    card['AvitoStatus'] = row[15].value
                else:
                    continue
                if row[16].value != None:
                    card['ContactMethod'] = row[16].value
                else:
                    continue
                if row[17].value != None:
                    card['Category'] = row[17].value
                else:
                    continue
                if row[18].value != None:
                    card['ListingFee'] = row[18].value
                else:
                    continue
                if row[19].value != None:
                    card['CompanyName'] = row[19].value
                else:
                    continue
                if row[20].value != None:
                    card['Quantity'] = row[20].value
                else:
                    continue
                if row[21].value != None:
                    card['TireType'] = row[21].value
                else:
                    continue
                if row[22].value != None:
                    card['TireAspectRatio'] = row[22].value
                else:
                    continue
                if row[23].value != None:
                    card['LoadIndex'] = row[23].value
                else:
                    continue
                if row[24].value != None:
                    card['DifferentWidthTires'] = row[24].value
                else:
                    continue
                if row[25].value != None:
                    card['SpeedIndex'] = row[25].value
                else:
                    continue
                if row[26].value != None:
                    card['RunFlat'] = row[26].value
                else:
                    continue
                if row[27].value != None:
                    card['Model'] = row[27].value
                else:
                    continue
                if row[28].value != None:
                    card['Brand'] = row[28].value
                else:
                    continue
                if row[29].value != None:
                    card['TireSectionWidth'] = row[29].value
                else:
                    continue
                if row[30].value != None:
                    card['VideoURL'] = row[30].value
                else:
                    continue
                # if row[31].value != None:
                #     card['DateBegin'] = row[31].value
                # else:
                #     card['DateBegin'] = None
                if row[32].value != None:
                    card['DateEnd'] = row[32].value
                else:
                    card['DateEnd'] = None
                if row[36].value != None:
                    card['main_img'] = row[36].value
                else:
                    card['main_img'] = None
                cards_list.append(card)
        
        if len(cards_list) > 0:
            return cards_list
        else:
            logging.error(f"<parse_tires> cards_list is empty")
            return 0
    else:
        raise Exception('Лист пружины содержит некорректный титул')
