import os
import io
import logging
from openpyxl import load_workbook

from .parse_rims import parse_rims
from .parse_springs import parse_springs
from .parse_tires import parse_tires
from .write_promo_xlsx import write_promo_xlsx
from info_bot import send_message
from config import ALLOY_RIMS_MOSCOW_XLSX, ALLOY_RIMS_KZN_DOROZH_XLSX, ALLOY_RIMS_SAMARA_XLSX,\
    FORGED_RIMS_MOSCOW_XLSX, FORGED_RIMS_KZN_DOROZH_XLSX, FORGED_RIMS_SAMARA_XLSX

def parse_xlsx(input_file_path: str, input_filename: str):
    """returns a dictionary with products cards (rims, tires, springs) collected from the xlsx file.

    :param input_file_path: a xlsx file path.
    :param input_filename: a name of input file.
    :return: a dictionary with the cards or zero when error occured.
    """

    forged_rims_cards = []
    alloy_rims_cards = []
    springs_cards = []
    tires_cards = []
    cards_dict = {}

    for root, dirs, files in os.walk(input_file_path):  
        for filename in files:
            if filename == input_filename:
                file_path = os.path.join(input_file_path, filename)
                if os.path.exists(file_path) and os.path.isfile(file_path):
                    with open(file_path, "rb") as f:
                        in_mem_file = io.BytesIO(f.read())
                    wb = load_workbook(in_mem_file, read_only=True)
                else:
                    send_message(f'ERROR\nfile path: {file_path} not exists')
                    logging.error(f'file path: {file_path} not exists')
                    return 0
                for sheetname in wb.sheetnames:
                    ws = wb[sheetname]
                    if sheetname == 'диски кованые':
                        logging.info('parse forged rims')
                        forged_rims_cards = parse_rims(ws=ws)
                        if forged_rims_cards == 0:
                            send_message('ERROR\nЛист кованные диски пуст!')
                    elif sheetname == 'диски литые':
                        logging.info('parse alloy rims')
                        alloy_rims_cards = parse_rims(ws=ws)
                        if alloy_rims_cards == 0:
                            send_message('ERROR\nЛист литые диски пуст!')
                    # Deprecated
                    # elif sheetname == 'пружины':
                    #     logging.info('parse springs')
                    #     springs_cards = parse_springs(ws=ws)
                        # if springs_cards == 0:
                        #     send_message('ERROR\nЛист пружины пуст!')
                    # Deprecated
                    # elif sheetname == 'шины':
                    #     logging.info('parse tires')
                    #     tires_cards = parse_tires(ws=ws)
                    #     if tires_cards == 0:
                    #         send_message('ERROR\nЛист шины пуст!')

                wb.close()
                
                if isinstance(alloy_rims_cards, list) and len(alloy_rims_cards) > 0:
                    cards_dict['alloy_rims_cards'] = alloy_rims_cards
                    logging.info(f'alloy cards: {len(alloy_rims_cards)}')
                    write_promo_xlsx(
                        save_path=ALLOY_RIMS_MOSCOW_XLSX,
                        region="Москва",
                        filter_key=["rest_count_msk", "rest_count_kzn_dorozh"],  
                        cards_list=alloy_rims_cards,
                        origin="price_origin_msk", 
                        promo="promo_price_msk",
                        simple_id=False,
                        discount_percent=5
                    )
                    write_promo_xlsx(
                        save_path=ALLOY_RIMS_KZN_DOROZH_XLSX,
                        region="Казань",
                        filter_key=["rest_count_kzn_dorozh", "rest_count_kazan"],  
                        cards_list=alloy_rims_cards,
                        simple_id=False,
                        discount_percent=5
                    )
                    write_promo_xlsx(
                        save_path=ALLOY_RIMS_SAMARA_XLSX,
                        region="Самара",
                        filter_key=["rest_count_samara", "rest_count_kzn_dorozh"],  
                        cards_list=alloy_rims_cards,
                        simple_id=True,
                        discount_percent=5
                    )

                if isinstance(forged_rims_cards, list) and len(forged_rims_cards) > 0:
                    cards_dict['forged_rims_cards'] = forged_rims_cards
                    logging.info(f'forged_rims_cards: {len(forged_rims_cards)}')

                    write_promo_xlsx(
                        save_path=FORGED_RIMS_MOSCOW_XLSX,
                        region="Москва",
                        filter_key=["rest_count_msk", "rest_count_kzn_dorozh"],  
                        cards_list=forged_rims_cards,
                        origin="price_origin_msk", 
                        promo="promo_price_msk",
                        simple_id=False,
                        discount_percent=0
                    )
                    write_promo_xlsx(
                        save_path=FORGED_RIMS_KZN_DOROZH_XLSX,
                        region="Казань",
                        filter_key=["rest_count_kzn_dorozh", "rest_count_kazan"],  
                        cards_list=forged_rims_cards,
                        simple_id=False,
                        discount_percent=0
                    )
                    write_promo_xlsx(
                        save_path=FORGED_RIMS_SAMARA_XLSX,
                        region="Самара",
                        filter_key=["rest_count_samara", "rest_count_kzn_dorozh"],  
                        cards_list=forged_rims_cards,
                        simple_id=True,
                        discount_percent=0
                    )
                # if isinstance(springs_cards, list) and len(springs_cards) > 0:
                #     cards_dict['springs_cards'] = springs_cards
                #     logging.info(f'springs_cards: {len(springs_cards)}')
                # if isinstance(tires_cards, list) and len(tires_cards) > 0:
                #     cards_dict['tires_cards'] = tires_cards
                #     logging.info(f'tires_cards: {len(tires_cards)}')
                
                return cards_dict

    return 0
