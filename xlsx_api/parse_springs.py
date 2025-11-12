from openpyxl import worksheet


def parse_springs(ws: worksheet):
    """returns a list with springs cards collected from the xlsx table.

    :param ws: a worksheet object.
    :return: a list with the cards or zero when error occured.
    """
    cards_list = []

    if ws[1][0].value == 'AvitoId' and ws[1][1].value == 'Id' and ws[1][2].value == 'ContactPhone'\
        and ws[1][3].value == 'ListingFee' and ws[1][4].value == 'AvitoDateEnd' and ws[1][5].value == 'ProductType'\
        and ws[1][6].value == 'Price' and ws[1][7].value == 'Description' and ws[1][8].value == 'CompanyName'\
        and ws[1][9].value == 'Title' and ws[1][10].value == 'ContactMethod' and ws[1][11].value == 'Category'\
        and ws[1][12].value == 'GoodsType' and ws[1][13].value == 'AdType' and ws[1][14].value == 'ImageUrls'\
        and ws[1][15].value == 'TypeId' and ws[1][16].value == 'Condition' and ws[1][17].value == 'EMail'\
        and ws[1][18].value == 'Address' and ws[1][19].value == 'AvitoStatus' and ws[1][20].value == 'Brand'\
        and ws[1][21].value == 'Model' and ws[1][22].value == 'Availability' and ws[1][23].value == 'Generation'\
        and ws[1][24].value == 'SparePartType' and ws[1][25].value == 'Make' and ws[1][26].value == 'Originality'\
        and ws[1][27].value == 'OEM':
        

        for index, row in enumerate(ws.rows):
            if row[1].value == 'Id':
                continue

            if row[1].value != None and index <= ws.max_row:
                card = {"Id" : row[1].value}
                if row[3].value != None:
                    card['ListingFee'] = row[3].value
                else:
                    continue
                if row[5].value != None:
                    card['ProductType'] = row[5].value
                else:
                    continue
                if row[6].value != None:
                    card['Price'] = row[6].value
                else:
                    continue 
                if row[7].value != None:
                    card['Description'] = row[7].value
                else:
                    continue
                if row[8].value != None:
                    card['CompanyName'] = row[8].value
                else:
                    continue
                if row[9].value != None:
                    card['Title'] = row[9].value
                else:
                    continue
                if row[10].value != None:
                    card['ContactMethod'] = row[10].value
                else:
                    continue
                if row[11].value != None:
                    card['Category'] = row[11].value
                else:
                    continue
                if row[12].value != None:
                    card['GoodsType'] = row[12].value
                else:
                    continue
                if row[13].value != None:
                    card['AdType'] = row[13].value
                else:
                    continue
                if row[14].value != None:
                    card['ImageUrls'] = row[14].value
                else:
                    continue 
                if row[15].value != None:
                    card['TypeId'] = row[15].value
                if row[16].value != None:
                    card['Condition'] = row[16].value
                else:
                    continue
                if row[19].value != None:
                    card['AvitoStatus'] = row[19].value
                else:
                    continue
                if row[20].value != None:
                    card['Brand'] = row[20].value
                else:
                    continue
                if row[21].value != None:
                    card['Model'] = row[21].value
                else:
                    continue
                if row[22].value != None:
                    card['Availability'] = row[22].value
                else:
                    continue
                
                card['Generation'] = row[23].value

                if row[24].value != None:
                    card['SparePartType'] = row[24].value
                else:
                    continue
                if row[25].value != None:
                    card['Make'] = row[25].value
                else:
                    continue
                if row[26].value != None:
                    card['Originality'] = row[26].value
                else:
                    continue
                card['OEM'] = row[27].value
                card['Modification'] = row[28].value
                card['BodyType'] = row[29].value
                card['Doors'] = row[30].value
                

                cards_list.append(card)
        
        if len(cards_list) > 0:
            return cards_list
        else:
            return 0
    else:
        raise Exception('Лист пружины содержит некорректный титул')
