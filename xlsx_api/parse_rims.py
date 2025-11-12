from openpyxl import worksheet

def try_parse_int(val):
    try:
        int_val = int(val)
        return int_val
    except Exception:
        return None

def try_parse_float(val):
    try:
        float_val = float(val)
        return float_val
    except Exception:
        return None

def check_pcd(in_pcd):
    float_pcd = try_parse_float(in_pcd)
    if float_pcd is not None:
        if float_pcd == 113:
            pcd = 112
        elif float_pcd == 112.5:
            pcd = 112
        elif float_pcd == 114.312:
            pcd = 114.3 
        else:
            pcd = float_pcd
        
        if isinstance(pcd, float) and pcd.is_integer():
            return int(pcd)
        else:
            return pcd
    else:
        return None

def parse_rims(ws: worksheet):
    """returns a list with rims cards collected from the xlsx table.

    :param ws: a worksheet object.
    :return: a list with the cards or zero when error occured.
    """
    cards_list = []
    title_list = []

    if ws[1][0].value == 'AvitoId' and ws[1][1].value == 'Id' and ws[1][2].value == 'Title'\
    and ws[1][3].value == 'Description' and ws[1][4].value == 'Price' and ws[1][5].value == 'RimDiameter'\
    and ws[1][6].value == 'RimBolts' and ws[1][7].value == 'RimBoltsDiameter' and ws[1][8].value == 'RimWidth'\
    and ws[1][9].value == 'RimOffset' and ws[1][10].value == 'RimDIA' and ws[1][11].value == 'RimType'\
    and ws[1][12].value == 'ProductType' and ws[1][13].value == 'ImageUrls' and ws[1][14].value == 'GoodsType'\
    and ws[1][15].value == 'AdType' and ws[1][19].value == 'TypeId' and ws[1][20].value == 'ManagerName'\
    and ws[1][21].value == 'Condition' and ws[1][22].value == 'AvitoStatus' and ws[1][23].value == 'ContactMethod'\
    and ws[1][24].value == 'Category' and ws[1][25].value == 'ListingFee' and ws[1][26].value == 'CompanyName'\
    and ws[1][27].value == 'DateBegin' and ws[1][16].value == 'Address' and ws[1][17].value == 'EMail'\
    and ws[1][18].value == 'ContactPhone' and ws[1][30].value == 'VideoUrl' and ws[1][31].value == 'RimBrand'\
    and ws[1][29].value == 'AdStatus':
        
        for index, row in enumerate(ws.rows):
            if row[1].value == 'Id':
                continue

            if row[1].value != None and index <= ws.max_row:
                card = {"Id" : row[1].value}
                if row[2].value != None and row[2].value.lower() not in title_list:
                    card['Title'] = row[2].value
                    title_list.append(row[2].value.lower())
                else:
                    continue
                if row[3].value != None:
                    card['Description'] = row[3].value
                else:
                    continue
                if row[4].value != None:
                    card['Price'] = row[4].value
                else:
                    continue
                if row[5].value != None:
                    card['RimDiameter'] = row[5].value
                else:
                    continue
                if row[6].value != None:
                    card['RimBolts'] = row[6].value
                else:
                    continue 
                if row[7].value != None and check_pcd(row[7].value):
                    card['RimBoltsDiameter'] = check_pcd(row[7].value)
                else:
                    continue
                if row[8].value != None:
                    card['RimWidth'] = row[8].value
                else:
                    continue
                if row[9].value != None:
                    card['RimOffset'] = row[9].value
                else:
                    continue
                if row[10].value != None:
                    card['RimDIA'] = row[10].value
                else:
                    continue
                if row[11].value != None:
                    card['RimType'] = row[11].value
                else:
                    continue
                if row[12].value != None:
                    card['ProductType'] = row[12].value
                else:
                    continue
                if row[13].value != None:
                    card['ImageUrls'] = row[13].value
                else:
                    continue
                if row[14].value != None:
                    card['GoodsType'] = row[14].value
                else:
                    continue 
                if row[15].value != None:
                    card['AdType'] = row[15].value
                else:
                    continue
                if row[19].value != None:
                    card['TypeId'] = row[19].value
                if row[20].value != None:
                    card['ManagerName'] = row[20].value
                else:
                    continue
                if row[21].value != None:
                    card['Condition'] = row[21].value
                else:
                    continue
                if row[22].value != None:
                    card['AvitoStatus'] = row[22].value
                else:
                    continue
                if row[23].value != None:
                    card['ContactMethod'] = row[23].value
                else:
                    continue
                if row[24].value != None:
                    card['Category'] = row[24].value
                else:
                    continue
                if row[25].value != None:
                    card['ListingFee'] = row[25].value
                else:
                    continue
                if row[26].value != None:
                    card['CompanyName'] = row[26].value
                else:
                    continue
                # if row[27].value != None:
                #     card['DateBegin'] = row[27].value
                # else:
                #     card['DateBegin'] = None
                if row[28].value != None:
                    card['DateEnd'] = row[28].value
                else:
                    card['DateEnd'] = None
                if row[29].value != None:
                    card['AdStatus'] = row[29].value
                else:
                    card['AdStatus'] = "Free"
                if row[30].value != None:
                    card['VideoUrl'] = row[30].value
                # if row[31].value != None:
                #     card['RimBrand'] = row[31].value
                # else:
                #     continue
                if row[32].value != None and row[32].value != '':
                    card['dorozh'] = row[32].value
                if row[33].value != None:
                    card['main_img'] = row[33].value
                else:
                    card['main_img'] = None
                if row[34].value != None and isinstance(row[34].value, str)\
                    and row[34].value.startswith('http'):
                    card['VideoFileURL'] = f'<![CDATA[{row[34].value}]]>'

                if len(row) > 35:
                    card['market_price'] = row[35].value if row[35].value else 0
                    card['promo_price'] = row[36].value if row[36].value else 0
                    card['price_origin_msk'] = row[37].value if row[37].value else 0
                    card['market_price_msk'] = row[38].value if row[38].value else 0
                    card['promo_price_msk'] = row[39].value if row[39].value else 0
                if len(row) > 40 and isinstance(row[40].value, str) and len(row[40].value) > 10:
                    card['specs'] = row[40].value
                    card['rest_count_kzn_dorozh'] = row[41].value
                    card['rest_count_kazan'] = row[42].value
                    card['rest_count_msk'] = row[43].value
                    card['rest_count_samara'] = row[44].value
                    card['rest_count_ufa'] = row[45].value
                    if row[46].value and row[46].value == "True":
                        card['portfolio'] = True
                    else:
                        card['portfolio'] = False
                    if row[47].value and row[47].value != "False":
                        card['RimBrand'] = row[47].value
                    else:
                        card['RimBrand'] = False
                    if row[49].value and row[49].value != "False":
                        card['RimModel'] = row[49].value
                    else:
                        card['RimModel'] = False
                    
                
                cards_list.append(card)

        if len(cards_list) > 0:
            return cards_list
        else:
            return 0
    else:
        raise Exception('Лист диски содержит некорректный титул')
