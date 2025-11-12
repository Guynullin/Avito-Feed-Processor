import os
import logging

from openpyxl import load_workbook


def get_input_tire_strings(input_file_path: str, input_filename: str):
    """collects tire titles from the 'Лист загрузки шин' sheet.

    :param input_file_path: the path to the input file.
    :param input_filename: the name of the input file.
    :return: the list with the tire titles or 0 if error occured.
    """
    string_list = []

    for root, dirs, files in os.walk(input_file_path):  
        for filename in files:
            if filename == input_filename:
                file_path = os.path.join(input_file_path, filename)
                if os.path.exists(file_path) and os.path.isfile(file_path):
                    wb = load_workbook(file_path, read_only=True)
                    if 'Лист загрузки шин' in wb.sheetnames:
                        ws_input = wb['Лист загрузки шин']
                        for index, row in enumerate(ws_input.rows):
                            if row[0].value != None and index <= ws_input.max_row:
                                string_list.append(row[0].value)
                    wb.close()

    if len(string_list) > 0:
        clear_string_list = list(set(string_list))
        w_h_list = []
        for input_string in clear_string_list:
            w_h = input_string.split("/")
            if len(w_h) == 2:
                flag = 0
                for item in w_h:
                    try:
                        int(item)
                    except:
                        flag += 1
                if flag > 0:
                    logging.error(f"<get_input_tire_strings> error string: {input_string}")
                    continue
                w_h_dict = {"width" : w_h[0].strip(), "height" : w_h[1].strip()}
                if w_h_dict in w_h_list:
                    continue
                else:
                    w_h_list.append(w_h_dict)
            else:
                logging.error(f"<get_input_tire_strings> len(w_h) != 2, error string: {input_string}")
                continue
        if len(w_h_list) > 0:
            return w_h_list
        else:
            logging.error(f"<get_input_tire_strings> w_h_list is empty, return 0")
            return 0
    else:
        logging.error(f"<get_input_tire_strings> string_list is empty, return 0")
        return 0
